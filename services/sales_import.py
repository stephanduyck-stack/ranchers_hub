"""Daily sales / financial data import.

Parses the Odoo Excel exports the finance team produces and upserts them into
the ``invoice`` table that all sales analytics read from. Designed to be run
repeatedly (daily) from the Admin screen:

* Rows are matched to an existing invoice by ``number`` and UPDATED in place;
  new numbers are inserted. Nothing is ever deleted, so a daily file may be
  the full year-to-date export or just the new days -- both work, and
  re-uploading the same file changes nothing.
* Two layouts are supported (chosen by the user on the upload screen):
    - "invoices"     -> "Invoice List Customer-Wise" export (positive amounts)
    - "credit_notes" -> "Credit Notes Customer List"   export (negative amounts)
* The itemized 'Journal Entry (account.move)' export (one row per invoice
  LINE) is auto-detected by its header row and routed to
  ``import_itemized`` regardless of the layout the user picked, so any of
  the three exports works from the same upload button.

Returns a summary dict: rows read, inserted, updated, customers matched.
"""
import re
import collections
import datetime

from openpyxl import load_workbook

from extensions import db
from models import Customer, Invoice


# Column maps (0-based) for each export layout.
LAYOUTS = {
    "invoices": {
        "number": 0, "customer": 2, "date": 4, "due": 5, "salesperson": 9,
        "untaxed": 11, "total": 12, "currency": 13, "payment": 14,
        "company_type": 7, "efris": 1,
    },
    "credit_notes": {
        "number": 0, "customer": 1, "date": 3, "due": 4, "salesperson": None,
        "untaxed": 6, "total": 7, "currency": None, "payment": 9,
        "company_type": None, "efris": 11,
    },
}


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _num(v):
    if isinstance(v, str):
        v = v.replace(",", "").strip()
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def _pick_sheet(wb):
    if "Sheet1" in wb.sheetnames:
        return wb["Sheet1"]
    return wb[wb.sheetnames[0]]


def _is_itemized(ws):
    """True when the sheet is the itemized 'Journal Entry (account.move)'
    export (one row per invoice LINE), recognized by its header row."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    cells = [str(c).strip() if c is not None else "" for c in header]
    return ("Invoice Partner Display Name" in cells
            or any("Invoice lines/Product" in c for c in cells))


def _check_layout(ws, layout):
    """Reject files whose header row does not match the chosen layout."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    cells = [str(c).strip() if c is not None else "" for c in header]
    first = cells[0] if cells else ""
    if layout == "invoices" and first and first.lower() != "number":
        raise ValueError(
            f"Unexpected file layout: first column header is '{first}', "
            "expected 'Number'. This screen expects the "
            "'Invoice List Customer-Wise' export.")


def import_file(path, layout="invoices"):
    """Upsert one Odoo export into the invoice table. Returns a summary dict."""
    if layout not in LAYOUTS:
        raise ValueError(f"Unknown layout '{layout}'")
    cmap = LAYOUTS[layout]
    wb = load_workbook(path, data_only=True)
    ws = _pick_sheet(wb)
    if _is_itemized(ws):
        # Wrong screen choice but a known export: route to the itemized
        # importer instead of failing (or worse, misreading the columns).
        return import_itemized(path)
    _check_layout(ws, layout)
    match = _build_matcher()

    def cell(row, key):
        idx = cmap[key]
        return row[idx] if idx is not None and idx < len(row) else None

    existing = {inv.number: inv for inv in db.session.scalars(db.select(Invoice))}
    cust_cache = {}
    read = inserted = updated = matched = dated = 0

    for r in ws.iter_rows(min_row=2, values_only=True):
        num = cell(r, "number")
        if num is None:
            continue
        num = str(num).strip()
        if not num:
            continue
        # skip year group-header rows like "2024 (123)"
        if "(" in num and num[:1].isdigit():
            continue
        read += 1

        cname = (cell(r, "customer") or "")
        cname = cname.strip() if isinstance(cname, str) else str(cname)
        if cname not in cust_cache:
            cust_cache[cname] = match(cname)
        cid = cust_cache[cname]
        if cid:
            matched += 1

        sp = cell(r, "salesperson")
        sp = (sp or "").strip() if isinstance(sp, str) else None
        if sp:
            sp = re.sub(r"\s*\(.*?\)\s*$", "", sp) or None

        currency = cell(r, "currency")
        currency = (currency or "UGX").strip() if isinstance(currency, str) else "UGX"
        payment = cell(r, "payment")
        payment = (payment or "").strip() if isinstance(payment, str) else None
        ctype = cell(r, "company_type")
        ctype = (ctype or "").strip() if isinstance(ctype, str) else None
        efris = cell(r, "efris")
        efris = str(efris) if efris not in (None, "") else None

        inv_date = _date(cell(r, "date"))
        if inv_date:
            dated += 1
        fields = dict(
            customer_id=cid, customer_name=cname, salesperson=sp,
            invoice_date=inv_date, due_date=_date(cell(r, "due")),
            untaxed=_num(cell(r, "untaxed")), total=_num(cell(r, "total")),
            currency=currency or "UGX", payment_status=payment or None,
            company_type=ctype or None, efris=efris)

        inv = existing.get(num)
        if inv:
            for k, v in fields.items():
                setattr(inv, k, v)
            updated += 1
        else:
            inv = Invoice(number=num, **fields)
            db.session.add(inv)
            existing[num] = inv
            inserted += 1

        if read % 5000 == 0:
            db.session.flush()

    # Invariant: a real invoice export always carries dates. Zero parsed
    # dates means a wrong or reshuffled layout -- abort instead of
    # committing rows the dated sales reports will never see.
    if read and dated == 0:
        db.session.rollback()
        raise ValueError(
            f"Read {read} rows but no invoice dates parsed. The file does "
            "not match the selected layout; nothing was imported.")

    db.session.commit()
    return {
        "layout": layout, "read": read, "inserted": inserted,
        "updated": updated, "customers": len(cust_cache),
        "matched_customers": sum(1 for v in cust_cache.values() if v),
        "matched_rows": matched,
    }


def _build_product_map():
    """Odoo product label -> catalogue product_id, from the curated
    sales_history linkage (link_history_products.py)."""
    from models import SalesHistory
    m = {}
    for label, pid in db.session.execute(
            db.select(SalesHistory.product, SalesHistory.product_id)
            .where(SalesHistory.product_id.isnot(None)).distinct()):
        if label and label not in m:
            m[label] = pid
    return m


def import_itemized(path):
    """Load the itemized 'Journal Entry (account.move)' export.

    One row per invoice LINE; rows with a Number start an invoice and
    continuation rows belong to the invoice above. Group/subtotal rows
    ("NAME (n)" or "27 Jul 2026 (70)") are skipped.

    Columns are located BY NAME (29 Jul 2026): finance re-picks columns in
    Odoo and the order shifts, so positional reads corrupted data once. Any
    column order works now. When the export carries 'Payment Status' and/or
    'Salesperson' columns those are applied to the invoices, including
    existing ones -- one file then keeps status current without the separate
    'Invoice List Customer-Wise' upload. Missing headers are created with
    customer matching (payment status defaults to 'Not Paid' when the file
    has no status column).

    * Lines: REPLACED per invoice on every run, so re-uploads are idempotent.
    * Amounts: journal-signed in the file (sales negative); sign is flipped
      once so lines are stored net positive.
    """
    import re as _re
    from models import InvoiceLine

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = _pick_sheet(wb)
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else ""
              for c in (next(rows, None) or ())]

    def col(name, required=False):
        if name in header:
            return header.index(name)
        if required:
            raise ValueError(
                f"Itemized invoice export is missing the '{name}' column. "
                "Re-export from Odoo with the standard columns and upload "
                "again. Nothing was imported.")
        return None

    i_name = col("Invoice Partner Display Name", required=True)
    i_num = col("Number", required=True)
    i_date = col("Invoice/Bill Date", required=True)
    i_prod = col("Invoice lines/Product", required=True)
    i_qty = col("Invoice lines/Quantity")
    i_amt = col("Invoice lines/Amount in Currency")
    i_total = col("Total Signed")
    if i_total is None:
        i_total = col("Total in Currency Signed")
    i_untaxed = col("Untaxed Amount Signed")
    i_status = col("Payment Status")
    i_sales = col("Salesperson")

    def cell(r, idx):
        return r[idx] if idx is not None and idx < len(r) else None

    _group_pat = _re.compile(r"\(\d+\)\s*$")

    cmatch = _build_matcher()
    pmap = _build_product_map()
    inv_by_number = {i.number: i for i in db.session.scalars(db.select(Invoice))}
    cust_cache = {}

    created_headers = touched = n_lines = matched_lines = 0
    statused = set()               # invoice ids whose status came from the file
    dated = 0
    cur = None                     # current Invoice object
    pending = []                   # lines for cur
    replaced = set()               # invoice ids whose lines were cleared this run

    def flush():
        nonlocal n_lines, matched_lines, pending
        if cur is None or not pending:
            pending = []
            return
        if cur.id not in replaced:
            db.session.query(InvoiceLine).filter_by(invoice_id=cur.id).delete(
                synchronize_session=False)
            replaced.add(cur.id)
        for prod, qty, amt in pending:
            pid = pmap.get(prod)
            if pid:
                matched_lines += 1
            db.session.add(InvoiceLine(
                invoice_id=cur.id, product_name=prod, product_id=pid,
                quantity=float(qty) if qty is not None else None,
                amount=-(_num(amt) or 0)))       # journal-signed -> net positive
            n_lines += 1
        pending = []

    for r in rows:
        num = cell(r, i_num)
        d = _date(cell(r, i_date))
        prod = cell(r, i_prod)
        if num is not None and str(num).strip():
            num = str(num).strip()
            # Group/subtotal rows ("NAME (12)", "27 Jul 2026 (70)") carry no
            # date and no product; year rows look like "2024 (123)".
            if _group_pat.search(num) and d is None and prod is None:
                continue
            flush()
            inv = inv_by_number.get(num)
            if d:
                dated += 1
            if inv is None:
                cname = (cell(r, i_name) or "")
                cname = cname.strip() if isinstance(cname, str) else str(cname)
                if cname not in cust_cache:
                    cust_cache[cname] = cmatch(cname)
                inv = Invoice(number=num, customer_id=cust_cache[cname],
                              customer_name=cname, invoice_date=d,
                              untaxed=_num(cell(r, i_untaxed)),
                              total=_num(cell(r, i_total)),
                              currency="UGX",
                              # NULL fails the dashboards' != 'Reversed'
                              # filter; default until a status column or the
                              # header export supplies the real status.
                              payment_status="Not Paid")
                db.session.add(inv)
                db.session.flush()
                inv_by_number[num] = inv
                created_headers += 1
            status = cell(r, i_status)
            if isinstance(status, str) and status.strip():
                inv.payment_status = status.strip()
                statused.add(id(inv))
            sales = cell(r, i_sales)
            if isinstance(sales, str) and sales.strip():
                inv.salesperson = _re.sub(r"\s*\(.*?\)\s*$", "",
                                          sales.strip()) or None
            cur = inv
            touched += 1
            if prod is not None:
                pending.append((str(prod).strip(), cell(r, i_qty), cell(r, i_amt)))
        elif prod is not None and cur is not None:
            pending.append((str(prod).strip(), cell(r, i_qty), cell(r, i_amt)))
        # else: customer subtotal / date group row -- ignore
        if n_lines and n_lines % 50000 == 0:
            db.session.flush()
    flush()
    # Invariant: a real itemized export carries invoice dates. Zero parsed
    # dates across all invoice rows means a wrong or reshuffled layout.
    if touched and dated == 0:
        db.session.rollback()
        raise ValueError(
            f"Read {touched} invoice rows but no invoice dates parsed. The "
            "file does not match the itemized export layout; nothing was "
            "imported.")
    db.session.commit()
    pct = matched_lines * 100 // max(n_lines, 1)
    tail = (f" Payment status set on {len(statused)} invoices from the file."
            if statused else
            " Upload the 'Invoice List Customer-Wise' export for the same "
            "days to fill in salesperson and payment status.")
    return {
        "layout": "itemized", "read": touched, "inserted": created_headers,
        "updated": touched - created_headers, "customers": len(cust_cache),
        "matched_customers": sum(1 for v in cust_cache.values() if v),
        "matched_rows": n_lines,
        "message": (f"Itemized export detected. {touched} invoices touched "
                    f"({created_headers} new), {n_lines} product lines loaded, "
                    f"{pct}% matched to the catalogue.{tail}"),
    }


# ---------------------------------------------------------------------------
# Daily finance uploads (27 Jul 2026): payments and receivable snapshot.
# Together with the invoice imports above these keep customer statements and
# the credit engine's outstanding balances current from three Odoo exports.
# ---------------------------------------------------------------------------

PAYMENTS_COLUMNS = ("Amount Company Currency Signed, Customer/Vendor, Date, "
                    "Journal, Number, Payment Method, Status")


def import_payments(path):
    """Load the 'Payments (account.payment)' export.

    Grouped export: date-group and customer-group rows carry no Number and
    are skipped; real payment rows have one. Upserted by payment number, so
    a full-history file and a days-only file both work and re-uploads change
    nothing.

    The header is validated before any row is read, and columns are located
    BY NAME, not position. Finance re-picking columns in Odoo once produced
    a file without the payment 'Number' column (29 Jul 2026): keyed on the
    neighbouring date column, every payment of a day collapsed into one
    record. A file without the Number column is now rejected outright with
    nothing imported."""
    from models import CustomerPayment

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = _pick_sheet(wb)
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else ""
              for c in (next(rows, None) or ())]

    def col(name, required=False):
        for i, h in enumerate(header):
            if h == name or h.startswith(name):
                return i
        if required:
            raise ValueError(
                f"Payments export is missing the '{name}' column. The Odoo "
                "export template changed. Re-export with the standard "
                f"columns ({PAYMENTS_COLUMNS}) and upload again. "
                "Nothing was imported.")
        return None

    i_num = col("Number", required=True)
    i_amt = col("Amount Company Currency", required=True)
    i_cust = col("Customer/Vendor", required=True)
    i_date = col("Date", required=True)
    i_journal = col("Journal")
    i_method = col("Payment Method")
    i_status = col("Status")
    # 'Number' must not resolve to 'Invoice Number' (the broken template).
    if header[i_num] != "Number":
        raise ValueError(
            "Payments export has no payment 'Number' column (found "
            f"'{header[i_num]}' instead). Re-export with the standard "
            f"columns ({PAYMENTS_COLUMNS}) and upload again. "
            "Nothing was imported.")

    def cell(r, idx):
        return r[idx] if idx is not None and idx < len(r) else None

    match = _build_matcher()
    existing = {p.number: p for p in db.session.scalars(db.select(CustomerPayment))}
    cust_cache = {}

    read = inserted = updated = 0
    seen_in_file = set()
    collisions = 0
    for r in rows:
        num = cell(r, i_num)
        if num is None or not str(num).strip():
            continue        # date-group / customer-group row
        if isinstance(num, (datetime.datetime, datetime.date)):
            db.session.rollback()
            raise ValueError(
                "Payments export carries dates in the 'Number' column; the "
                "file does not match the standard template "
                f"({PAYMENTS_COLUMNS}). Nothing was imported.")
        num = str(num).strip()
        read += 1
        if num in seen_in_file:
            collisions += 1
        seen_in_file.add(num)
        cname = cell(r, i_cust) or ""
        cname = cname.strip() if isinstance(cname, str) else str(cname)
        if cname not in cust_cache:
            cust_cache[cname] = match(cname)
        journal = cell(r, i_journal)
        method = cell(r, i_method)
        status = cell(r, i_status)
        fields = dict(
            customer_id=cust_cache[cname], customer_name=cname,
            payment_date=_date(cell(r, i_date)), amount=_num(cell(r, i_amt)),
            journal=journal if isinstance(journal, str) and journal else None,
            method=method if isinstance(method, str) and method else None,
            status=status if isinstance(status, str) and status else None)
        pay = existing.get(num)
        if pay:
            for k, v in fields.items():
                setattr(pay, k, v)
            updated += 1
        else:
            pay = CustomerPayment(number=num, **fields)
            db.session.add(pay)
            existing[num] = pay
            inserted += 1
        if read % 5000 == 0:
            db.session.flush()

    db.session.commit()
    msg = (f"Payments export detected. {read} payments read: "
           f"{inserted} new, {updated} updated.")
    if collisions:
        msg += (f" WARNING: {collisions} rows repeated a payment number "
                "already in the file and overwrote the earlier row. Check "
                "the export for duplicated payment numbers.")
    return {
        "layout": "payments", "read": read, "inserted": inserted,
        "updated": updated, "customers": len(cust_cache),
        "matched_customers": sum(1 for v in cust_cache.values() if v),
        "matched_rows": read,
        "message": msg,
    }


def import_receivables(path):
    """Load the 'Contact (res.partner)' export: per-customer Total Receivable.

    Stores the receivable as the customer's fresh balance snapshot
    (odoo_receivable, dated today). The credit engine reads outstanding as
    snapshot + unpaid invoices imported after the snapshot date, so this file
    is what actually resets customer balances each day. Rows that match the
    same app customer are summed (e.g. several POS partners on one account);
    negative receivables (customer credit) are kept."""
    import datetime as _dt
    from models import Customer

    match = _build_matcher()
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = _pick_sheet(wb)
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header

    read = 0
    per_cust = {}
    unmatched = []
    for r in rows:
        name = (r[0] or "")
        name = name.strip() if isinstance(name, str) else str(name)
        if not name:
            continue
        amount = _num(r[2] if len(r) > 2 else None)
        if amount is None:
            continue
        read += 1
        cid = match(name)
        if cid is None:
            unmatched.append(name)
            continue
        per_cust[cid] = per_cust.get(cid, 0.0) + amount

    today = _dt.date.today()
    for cid, amount in per_cust.items():
        c = db.session.get(Customer, cid)
        if c is not None:
            c.odoo_receivable = round(amount, 2)
            c.odoo_receivable_at = today
    db.session.commit()
    return {
        "layout": "receivables", "read": read, "inserted": 0,
        "updated": len(per_cust), "customers": read,
        "matched_customers": len(per_cust), "matched_rows": len(per_cust),
        "unmatched_names": unmatched,
        "message": (f"Receivable snapshot detected. {len(per_cust)} customer "
                    f"balances reset as of {today:%d %b %Y}; "
                    f"{len(unmatched)} export names had no matching account."),
    }


def detect_layout(path):
    """Identify which of the daily Odoo exports a workbook is, from its
    header row. Returns 'itemized' | 'payments' | 'receivables' |
    'invoices' | 'credit_notes', or raises ValueError."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = _pick_sheet(wb)
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    cells = [str(c).strip() if c is not None else "" for c in header]
    joined = " | ".join(cells)
    if "Invoice Partner Display Name" in cells or "Invoice lines/Product" in joined:
        return "itemized"
    if cells and cells[0].startswith("Amount Company Currency") and "Customer/Vendor" in cells:
        return "payments"
    if "Translated Display Name" in cells and "Total Receivable" in cells:
        return "receivables"
    if cells and cells[0].lower() == "number":
        return "invoices" if any("Salesperson" in c for c in cells) else "credit_notes"
    # Monthly product pivot: no header names, but row 2 carries month labels.
    from services.product_import import _month
    row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None) or ()
    if any(_month(v) for v in row2 if v is not None):
        return "product_pivot"
    raise ValueError(
        "Unrecognized file. Expected one of the Odoo exports: itemized "
        "invoices (Journal Entry), payments (account.payment), customer "
        "receivables (res.partner), the invoice list or the monthly product "
        f"pivot. Headers found: {joined[:200]}")


def import_auto(path):
    """Route one uploaded workbook to the right importer by its header row."""
    layout = detect_layout(path)
    if layout == "itemized":
        return import_itemized(path)
    if layout == "payments":
        return import_payments(path)
    if layout == "receivables":
        return import_receivables(path)
    if layout == "product_pivot":
        from services import product_import as pi
        r = pi.import_monthly_pivot(path)
        return {
            "layout": "product_pivot", "read": r["rows"], "inserted": r["rows"],
            "updated": 0, "customers": 0, "matched_customers": 0,
            "matched_rows": r["rows"],
            "message": (f"Product pivot detected. {r['rows']} rows across "
                        f"{r['months']} month(s) ({r['span']}); "
                        f"{r['linked_pct']}% of revenue linked to the catalogue."),
        }
    return import_file(path, layout)
